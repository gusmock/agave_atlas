from __future__ import annotations

import csv
import re
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from . import db


HEADER_SENTINELS = {"Publication numbers", "Title", "Abstract"}
BASE_FIELDS_END = 22


@dataclass
class ParsedColumn:
    index: int
    letter: str
    header: str
    group_name: str | None
    subgroup_name: str | None
    slug: str


def slugify(value: str, fallback: str) -> str:
    text = unicodedata.normalize("NFKD", value or "").encode("ascii", "ignore").decode()
    text = re.sub(r"[^a-zA-Z0-9]+", "_", text).strip("_").lower()
    return text or fallback


def clean(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).replace("\r\n", "\n").strip()
    return text if text else None


def fill_merged_value(ws, row: int, col: int, *, exact_merge_row: bool = False) -> Any:
    value = ws.cell(row, col).value
    if value not in (None, ""):
        return value
    for merged in ws.merged_cells.ranges:
        if exact_merge_row and merged.min_row != row:
            continue
        if merged.min_row <= row <= merged.max_row and merged.min_col <= col <= merged.max_col:
            return ws.cell(merged.min_row, merged.min_col).value
    return None


def find_header_row(ws) -> int:
    for row_idx in range(1, min(ws.max_row, 30) + 1):
        values = {clean(cell.value) for cell in ws[row_idx]}
        if HEADER_SENTINELS.issubset(values):
            return row_idx
    raise ValueError("Nao foi possivel localizar a linha de cabecalho da planilha.")


def useful_max_column(ws, header_row: int) -> int:
    max_col = 1
    for row_idx in range(1, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            value = ws.cell(row_idx, col_idx).value
            if value not in (None, ""):
                max_col = max(max_col, col_idx)
    return max_col


def parse_columns(ws, header_row: int, max_col: int) -> list[ParsedColumn]:
    columns: list[ParsedColumn] = []
    for col_idx in range(1, max_col + 1):
        raw_header = clean(ws.cell(header_row, col_idx).value)
        group = clean(fill_merged_value(ws, header_row - 2, col_idx))
        subgroup = clean(fill_merged_value(ws, header_row - 1, col_idx, exact_merge_row=True))

        if raw_header and raw_header.startswith("="):
            raw_header = raw_header[1:]
        if raw_header == "#REF!":
            raw_header = None

        if col_idx <= BASE_FIELDS_END:
            label = raw_header or f"Campo {col_idx}"
            group = None
            subgroup = None
        else:
            if not group and subgroup:
                group = "Pré tratamento"
            label_parts = [part for part in (group, subgroup, raw_header) if part]
            label = " / ".join(label_parts) if label_parts else f"Campo analisado {col_idx}"

        slug = slugify(label, f"col_{col_idx}")
        columns.append(
            ParsedColumn(
                index=col_idx,
                letter=get_column_letter(col_idx),
                header=label,
                group_name=group,
                subgroup_name=subgroup,
                slug=f"{slug}_{col_idx}",
            )
        )
    return columns


def row_is_empty(values: list[Any]) -> bool:
    return all(value in (None, "") for value in values)


def infer_primary_identifier(publication_numbers: str | None) -> str | None:
    if not publication_numbers:
        return None
    for part in re.split(r"[\n;,\s]+", publication_numbers):
        part = part.strip()
        if part:
            return part
    return publication_numbers.strip()


def is_analysis_column(column: ParsedColumn) -> bool:
    return column.index > BASE_FIELDS_END


def as_flag(value: str | None) -> int:
    if not value:
        return 0
    normalized = value.strip().lower()
    return 1 if normalized in {"1", "sim", "yes", "x", "true"} else 0


def import_xlsx(path: Path, db_path: Path | None = None) -> dict[str, Any]:
    conn = db.connect(db_path)
    try:
        db.init_db(conn)
        workbook = load_workbook(path, read_only=False, data_only=True, keep_vba=False)
        ws = workbook[workbook.sheetnames[0]]
        header_row = find_header_row(ws)
        max_col = useful_max_column(ws, header_row)
        columns = parse_columns(ws, header_row, max_col)

        run_id = conn.execute(
            """
            INSERT INTO import_runs(source_path, sheet_name, header_row, data_start_row, created_at)
            VALUES (?, ?, ?, ?, ?)
            """,
            (str(path), ws.title, header_row, header_row + 1, db.utcnow()),
        ).lastrowid

        source_column_ids: dict[int, int] = {}
        for column in columns:
            source_column_ids[column.index] = conn.execute(
                """
                INSERT INTO source_columns(import_run_id, sheet_name, column_index, letter,
                                           header, group_name, subgroup_name, slug)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    run_id,
                    ws.title,
                    column.index,
                    column.letter,
                    column.header,
                    column.group_name,
                    column.subgroup_name,
                    column.slug,
                ),
            ).lastrowid

        imported = 0
        field_by_header = db.DOCUMENT_FIELDS
        for excel_row_idx in range(header_row + 1, ws.max_row + 1):
            values = [ws.cell(excel_row_idx, col.index).value for col in columns]
            if row_is_empty(values):
                continue

            raw_by_label: dict[str, Any] = {}
            doc_values: dict[str, Any] = {"import_run_id": run_id, "document_type": "patente"}
            for column, value in zip(columns, values):
                text = clean(value)
                raw_by_label[column.header] = text
                target = field_by_header.get(column.header)
                if target:
                    doc_values[target] = text
            doc_values["primary_identifier"] = infer_primary_identifier(
                doc_values.get("publication_numbers")
            )
            doc_id = db.insert_document(conn, doc_values, raw=raw_by_label)

            for column, value in zip(columns, values):
                text = clean(value)
                if text is None:
                    continue
                source_column_id = source_column_ids[column.index]
                conn.execute(
                    """
                    INSERT INTO document_field_values(document_id, source_column_id, value)
                    VALUES (?, ?, ?)
                    """,
                    (doc_id, source_column_id, text),
                )
                if is_analysis_column(column):
                    conn.execute(
                        """
                        INSERT INTO document_categories(document_id, group_name, subgroup_name,
                                                        field_label, value_text, flag_value,
                                                        source_column_id)
                        VALUES (?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            doc_id,
                            column.group_name or "Analise",
                            column.subgroup_name,
                            column.header,
                            text,
                            as_flag(text),
                            source_column_id,
                        ),
                    )
            imported += 1

        conn.execute(
            "UPDATE import_runs SET rows_imported = ? WHERE id = ?", (imported, run_id)
        )
        db.rebuild_fts(conn)
        conn.commit()
        return {
            "run_id": run_id,
            "sheet_name": ws.title,
            "header_row": header_row,
            "rows_imported": imported,
            "columns_imported": len(columns),
            "db_path": str(db_path or db.settings.resolved_db_path()),
        }
    finally:
        conn.close()


def import_csv(path: Path, db_path: Path | None = None) -> dict[str, Any]:
    conn = db.connect(db_path)
    try:
        db.init_db(conn)
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            headers = reader.fieldnames or []
            run_id = conn.execute(
                """
                INSERT INTO import_runs(source_path, sheet_name, header_row, data_start_row, created_at)
                VALUES (?, ?, ?, ?, ?)
                """,
                (str(path), path.name, 1, 2, db.utcnow()),
            ).lastrowid
            source_column_ids = {}
            for idx, header in enumerate(headers, 1):
                label = header.strip() or f"Campo {idx}"
                source_column_ids[label] = conn.execute(
                    """
                    INSERT INTO source_columns(import_run_id, sheet_name, column_index, letter,
                                               header, group_name, subgroup_name, slug)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (run_id, path.name, idx, str(idx), label, None, None, slugify(label, f"col_{idx}")),
                ).lastrowid
            imported = 0
            for row in reader:
                doc_values = {"import_run_id": run_id, "document_type": "documento"}
                for header, target in db.DOCUMENT_FIELDS.items():
                    if header in row:
                        doc_values[target] = clean(row[header])
                doc_values["primary_identifier"] = infer_primary_identifier(
                    doc_values.get("publication_numbers")
                )
                doc_id = db.insert_document(conn, doc_values, raw=row)
                for header, value in row.items():
                    text = clean(value)
                    if text is None:
                        continue
                    conn.execute(
                        """
                        INSERT INTO document_field_values(document_id, source_column_id, value)
                        VALUES (?, ?, ?)
                        """,
                        (doc_id, source_column_ids[header], text),
                    )
                imported += 1
            conn.execute("UPDATE import_runs SET rows_imported = ? WHERE id = ?", (imported, run_id))
            db.rebuild_fts(conn)
            conn.commit()
            return {
                "run_id": run_id,
                "sheet_name": path.name,
                "header_row": 1,
                "rows_imported": imported,
                "columns_imported": len(headers),
                "db_path": str(db_path or db.settings.resolved_db_path()),
            }
    finally:
        conn.close()


def import_spreadsheet(path: str | Path, db_path: str | Path | None = None) -> dict[str, Any]:
    source = Path(path)
    target = Path(db_path) if db_path else None
    suffix = source.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        return import_xlsx(source, target)
    if suffix == ".csv":
        return import_csv(source, target)
    raise ValueError(f"Formato nao suportado: {source.suffix}")
